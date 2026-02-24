"""
Campaign Optimization Engine
Digital Turbine Preload Campaign Optimizer
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO


# ─── CONSTANTS ────────────────────────────────────────────────────────────────

EXCLUDED_SITE_PATTERNS = r'OM.?Push|OM_PUSH|Notif'

FILLS = {
    'green':  'C6EFCE',
    'yellow': 'FFEB9C',
    'orange': 'FFCC99',
    'red':    'FFC7CE',
    'header': '1F3864',
    'cap':    'DAE3F3',
}


# ─── STEP 1: DATA LOADING & PREP ──────────────────────────────────────────────

def load_internal(filepath):
    return pd.read_excel(filepath)


def load_advertiser(filepath):
    return pd.read_csv(filepath)


def exclude_site_types(df):
    """Remove OM Push, Notification, and similar site types."""
    mask = (
        df['siteName'].str.contains(EXCLUDED_SITE_PATTERNS, case=False, na=False, regex=True) |
        df['campaignName'].str.contains(EXCLUDED_SITE_PATTERNS, case=False, na=False, regex=True)
    )
    excluded = mask.sum()
    return df[~mask].copy(), excluded


def build_key(df, campaign_col, site_col):
    """Create Key = campaignName_siteId."""
    return df[campaign_col].str.strip() + '_' + df[site_col].astype(int).astype(str)


def parse_pct(val):
    """Parse '5.9%' or 0.059 → float decimal."""
    if pd.isna(val):
        return None
    s = str(val).replace('%', '').strip()
    try:
        v = float(s)
        # If value looks like a whole percentage (e.g. 5.9), divide by 100
        return v / 100.0 if v > 1.0 else v
    except:
        return None


def merge_advertiser_data(internal, advertiser, kpi_col_idx_d7, kpi_col_idx_d2nd):
    """
    Merge ROI columns from advertiser file into internal file.
    kpi_col_idx_d7: 0-based column index for D7 KPI (e.g. 8 for column I)
    kpi_col_idx_d2nd: 0-based column index for D2nd KPI (e.g. 10 for column K)
    Returns (merged_df, d7_col_name, d2nd_col_name)
    """
    # Detect column names
    d7_col_raw   = advertiser.columns[kpi_col_idx_d7]
    d2nd_col_raw = advertiser.columns[kpi_col_idx_d2nd]

    # Build friendly label from column name (e.g. "Full ROAS D30" → "ROI D30")
    def label(col):
        c = col.strip()
        for prefix in ['Full ROAS ', 'ROAS ', 'Full Roas ', 'ROI ']:
            if prefix.lower() in c.lower():
                period = c.split()[-1]  # e.g. D7, D30, D14
                return f'ROI {period}'
        return f'ROI {c}'

    d7_label   = label(d7_col_raw)
    d2nd_label = label(d2nd_col_raw)

    # Parse percentages
    advertiser = advertiser.copy()
    advertiser[d7_label]   = advertiser[d7_col_raw].apply(parse_pct)
    advertiser[d2nd_label] = advertiser[d2nd_col_raw].apply(parse_pct)

    # Detect key columns in advertiser file
    # Common patterns: 'Campaign Name', 'Campaign', 'campaign_name'
    camp_col = next((c for c in advertiser.columns if 'campaign' in c.lower() and 'name' in c.lower()), None)
    if camp_col is None:
        camp_col = next((c for c in advertiser.columns if 'campaign' in c.lower()), None)
    site_col = next((c for c in advertiser.columns if 'site' in c.lower() and 'id' in c.lower()), None)

    advertiser['Key'] = build_key(advertiser, camp_col, site_col)
    internal['Key']   = build_key(internal, 'campaignName', 'siteId')

    lookup = advertiser[['Key', d7_label, d2nd_label]].drop_duplicates('Key')
    merged = internal.merge(lookup, on='Key', how='left')

    return merged, d7_label, d2nd_label


def clean_nulls(df, d7_col, d2nd_col):
    """Drop rows with null in critical columns after merge."""
    before = len(df)
    df = df.dropna(subset=[d7_col, d2nd_col, 'maxPreloads', 'fillRate'])
    dropped = before - len(df)
    return df.reset_index(drop=True), dropped


# ─── STEP 2: SEGMENTATION ─────────────────────────────────────────────────────

def weighted_score(roi_d7, roi_d2nd, w_d7=0.80, w_d2nd=0.20):
    return w_d7 * roi_d7 + w_d2nd * roi_d2nd


def kpi_target(kpi_d7, kpi_d2nd, w_d7=0.80, w_d2nd=0.20):
    return w_d7 * kpi_d7 + w_d2nd * kpi_d2nd


def segment_row(roi_d7, roi_d2nd, kpi_d7, kpi_d2nd):
    score  = weighted_score(roi_d7, roi_d2nd)
    target = kpi_target(kpi_d7, kpi_d2nd)
    if score >= target:
        return 'green'
    if roi_d7 == 0 and roi_d2nd == 0:
        return 'red'
    pct_below = (target - score) / target
    if pct_below <= 0.50:
        return 'yellow'
    if pct_below < 1.0:
        return 'orange'
    return 'red'


def segment_single(val, kpi):
    """Segment a single ROI value against its own KPI (used for poor progression check)."""
    if val >= kpi:      return 'green'
    if val == 0:        return 'red'
    pct_below = (kpi - val) / kpi
    if pct_below <= 0.50: return 'yellow'
    if pct_below < 1.0:   return 'orange'
    return 'red'


def get_progression(roi_d7, roi_d2nd):
    """Detect good/poor/flat progression. Whale protection: D7=0 → flat."""
    if roi_d7 == 0:
        return 'flat'
    if roi_d2nd > roi_d7:
        return 'good'
    if roi_d2nd < roi_d7:
        return 'poor'
    return 'flat'


def add_segments(df, d7_col, d2nd_col, kpi_d7, kpi_d2nd):
    df = df.copy()
    df['segment']    = df.apply(lambda r: segment_row(r[d7_col], r[d2nd_col], kpi_d7, kpi_d2nd), axis=1)
    df['progression'] = df.apply(lambda r: get_progression(r[d7_col], r[d2nd_col]), axis=1)
    return df


# ─── STEP 3: DISCARD LOGIC ────────────────────────────────────────────────────

def should_discard(row):
    is_green   = row['segment'] == 'green'
    installs   = row['installs'] if pd.notna(row['installs']) else 0
    spend      = row['spend']
    preloads   = row['preloads']
    progression = row['progression']
    fill_rate  = row['fillRate'] if pd.notna(row['fillRate']) else 0
    d7         = row.get('_roi_d7', 0)

    # Green exception: green + 5+ installs overrides spend AND preloads thresholds
    if is_green and installs >= 5:
        return False

    # Good progression exception: overrides discard for yellow/orange
    if progression == 'good' and fill_rate < 0.60 and d7 > 0 and spend >= 100:
        return False

    if spend < 100:    return True
    if preloads < 100: return True
    if str(row['status']).lower() == 'paused' and not is_green: return True

    return False


# ─── STEP 4: DAILY CAP LOGIC ──────────────────────────────────────────────────

def get_daily_cap_suggestion(row, d7_col, d2nd_col):
    if row['spend'] <= 1000:
        return None

    floor    = row['effectiveBidFloor'] if pd.notna(row['effectiveBidFloor']) else None
    bid      = row['bidRate']
    at_floor = floor is not None and bid <= floor
    has_perf = row[d7_col] > 0 or row[d2nd_col] > 0
    no_perf  = row[d7_col] == 0 and row[d2nd_col] == 0

    if no_perf:
        return 'Suggest pause'

    if at_floor and has_perf:
        daily = round((row['spend'] / 30) * 0.50, 2)
        return f'Add daily cap ${daily:.2f}'

    return None


# ─── STEP 5: BID OPTIMIZATION ─────────────────────────────────────────────────

def optimize_bid(row, d7_col, d2nd_col, kpi_d7, kpi_d2nd):
    floor      = row['effectiveBidFloor'] if pd.notna(row['effectiveBidFloor']) else None
    bid        = row['bidRate']
    high_tier  = row['highTier'] if pd.notna(row['highTier']) else None
    fill_rate  = row['fillRate'] if pd.notna(row['fillRate']) else 0
    installs   = row['installs'] if pd.notna(row['installs']) else 0
    seg        = row['segment']
    prog       = row['progression']
    roi_d7     = row[d7_col]
    roi_d2nd   = row[d2nd_col]
    spend      = row['spend']
    at_floor   = floor is not None and bid <= floor

    score      = weighted_score(roi_d7, roi_d2nd)
    target     = kpi_target(kpi_d7, kpi_d2nd)
    pct_above  = (score - target) / target if target > 0 else 0
    pct_below  = (target - score) / target if target > 0 else 1

    def apply_floor(new_bid):
        if floor is not None and new_bid < floor:
            return round(floor, 2), 'Meet bid floor'
        return round(new_bid, 2), None

    def apply_high_tier(new_bid, pct):
        """Cap new bid at highTier. If already above tier + fill < 70%, allow up to 15%."""
        if high_tier is None:
            return round(new_bid, 2), f'Increase bid {int(pct*100)}%'
        if bid > high_tier and fill_rate < 0.70:
            capped = round(bid * 1.15, 2)
            return capped, 'Increase bid 15%'
        capped = round(min(new_bid, high_tier), 2)
        return capped, f'Increase bid {int(pct*100)}%'

    # ── Discard check ────────────────────────────────────────────────────────
    if row['discard']:
        return None, None

    # ── Already at/below floor → no bid suggestion ───────────────────────────
    if at_floor:
        return None, None

    # ── GOOD PROGRESSION (overrides segment) ─────────────────────────────────
    if prog == 'good' and fill_rate < 0.60 and roi_d7 > 0:
        ratio        = roi_d2nd / roi_d7 if roi_d7 > 0 else 1
        increase_pct = 0.15 if ratio >= 2.0 else 0.10
        new_bid      = round(bid * (1 + increase_pct), 2)
        new_bid, action = apply_high_tier(new_bid, increase_pct)
        floored, floor_action = apply_floor(new_bid)
        if floor_action:
            return floor_action, floored
        return action, new_bid

    # ── POOR PROGRESSION ─────────────────────────────────────────────────────
    if prog == 'poor':
        if spend < 100:
            return None, None
        seg_d2nd = segment_single(roi_d2nd, kpi_d2nd)
        if seg_d2nd in ('yellow', 'orange', 'red'):
            new_bid = round(bid * 0.90, 2)
            floored, floor_action = apply_floor(new_bid)
            if floor_action:
                return floor_action, floored
            return 'Decrease bid 10%', new_bid
        return None, None  # D2nd green → wait and see

    # ── GREEN ────────────────────────────────────────────────────────────────
    if seg == 'green':
        if installs < 5:
            return None, None

        if fill_rate > 0.80:
            # Cap at 15%
            new_bid = round(bid * 1.15, 2)
            if high_tier is not None:
                new_bid = round(min(new_bid, high_tier), 2)
            floored, floor_action = apply_floor(new_bid)
            return (floor_action or 'Increase bid 15%'), (floored if floor_action else new_bid)

        # Fill 60–80%: still cap at 15%
        if fill_rate > 0.60:
            increase_pct = 0.15
        else:
            # Fill ≤ 60%: normal 10/20/30%
            increase_pct = 0.10 if pct_above <= 0.25 else (0.20 if pct_above <= 0.50 else 0.30)

        new_bid = round(bid * (1 + increase_pct), 2)
        new_bid, action = apply_high_tier(new_bid, increase_pct)
        floored, floor_action = apply_floor(new_bid)
        if floor_action:
            return floor_action, floored
        return action, new_bid

    # ── YELLOW ───────────────────────────────────────────────────────────────
    if seg == 'yellow':
        decrease_pct = 0.10 if pct_below <= 0.25 else 0.15
        new_bid = round(bid * (1 - decrease_pct), 2)
        floored, floor_action = apply_floor(new_bid)
        if floor_action:
            return floor_action, floored
        return f'Decrease bid {int(decrease_pct*100)}%', new_bid

    # ── ORANGE ───────────────────────────────────────────────────────────────
    if seg == 'orange':
        decrease_pct = 0.20 if pct_below <= 0.75 else 0.25
        new_bid = round(bid * (1 - decrease_pct), 2)
        floored, floor_action = apply_floor(new_bid)
        if floor_action:
            return floor_action, floored
        return f'Decrease bid {int(decrease_pct*100)}%', new_bid

    # ── RED ──────────────────────────────────────────────────────────────────
    new_bid = round(bid * 0.70, 2)
    floored, floor_action = apply_floor(new_bid)
    if floor_action:
        return floor_action, floored
    return 'Decrease bid 30%', new_bid


# ─── MAIN PIPELINE ────────────────────────────────────────────────────────────

def run_optimization(
    internal_path,
    advertiser_path,
    kpi_col_idx_d7,    # 0-based index, e.g. 8 for column I
    kpi_col_idx_d2nd,  # 0-based index, e.g. 10 for column K
    kpi_d7,            # e.g. 0.0336 for 3.36%
    kpi_d2nd,          # e.g. 0.1336 for 13.36%
):
    """
    Full optimization pipeline.
    Returns (output_bytes, summary_dict)
    output_bytes: Excel file as BytesIO ready for download
    summary_dict: stats about the run
    """
    # Load
    internal   = load_internal(internal_path)
    advertiser = load_advertiser(advertiser_path)

    # Exclude site types
    internal, excluded_count = exclude_site_types(internal)

    # Merge advertiser KPIs
    internal, d7_col, d2nd_col = merge_advertiser_data(
        internal, advertiser, kpi_col_idx_d7, kpi_col_idx_d2nd
    )

    # Clean nulls
    internal, dropped_count = clean_nulls(internal, d7_col, d2nd_col)

    # Segmentation + progression
    internal = add_segments(internal, d7_col, d2nd_col, kpi_d7, kpi_d2nd)

    # Store roi_d7 reference for discard helper
    internal['_roi_d7'] = internal[d7_col]

    # Discard
    internal['discard'] = internal.apply(should_discard, axis=1)

    # Daily cap
    internal['Daily Cap Suggestion'] = internal.apply(
        lambda r: get_daily_cap_suggestion(r, d7_col, d2nd_col), axis=1
    )

    # Bid optimization
    results = internal.apply(
        lambda r: pd.Series(optimize_bid(r, d7_col, d2nd_col, kpi_d7, kpi_d2nd)),
        axis=1
    )
    results.columns = ['Action', 'Recommended bid']
    internal['Action']          = results['Action']
    internal['Recommended bid'] = results['Recommended bid']

    # Output columns
    out_cols = [
        'Key', 'campaignId', 'campaignName', 'siteId', 'siteName', 'status',
        'spend', 'preloads', 'maxPreloads', 'fillRate', 'installs', 'cvr',
        'ecpp', 'ecpi', 'bidFloorGroupName', 'effectiveBidFloor', 'bidRate',
        'dailyCap', 'lowTier', 'midTier', 'highTier',
        d7_col, d2nd_col, 'Action', 'Recommended bid', 'Daily Cap Suggestion'
    ]
    # Only keep columns that exist
    out_cols = [c for c in out_cols if c in internal.columns]
    output   = internal[out_cols].copy()

    # Build Excel
    excel_bytes = build_excel(output, internal, d7_col, d2nd_col, out_cols)

    # Summary
    summary = {
        'total_rows':       len(output),
        'excluded':         excluded_count,
        'dropped_nulls':    dropped_count,
        'actioned':         int(output['Action'].notna().sum()),
        'disregarded':      int(output['Action'].isna().sum()),
        'daily_cap':        int(output['Daily Cap Suggestion'].notna().sum()),
        'action_breakdown': output['Action'].value_counts().to_dict(),
        'segment_breakdown': internal['segment'].value_counts().to_dict(),
        'd7_col':           d7_col,
        'd2nd_col':         d2nd_col,
    }

    return excel_bytes, summary


# ─── EXCEL BUILDER ────────────────────────────────────────────────────────────

def build_excel(output, internal, d7_col, d2nd_col, out_cols):
    seg_map     = dict(zip(internal['Key'], internal['segment']))
    discard_map = dict(zip(internal['Key'], internal['discard']))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Optimization'

    # Header
    for col_idx, h in enumerate(out_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = PatternFill('solid', start_color=FILLS['header'])
        cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, (_, row) in enumerate(output.iterrows(), 2):
        key      = row['Key']
        seg      = seg_map.get(key, '')
        disc     = discard_map.get(key, True)
        row_fill = PatternFill('solid', start_color=FILLS[seg]) if not disc and seg in FILLS else None

        for col_idx, col_name in enumerate(out_cols, 1):
            val = row[col_name]
            if pd.isna(val): val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name='Arial', size=9)
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # ROI columns — colour with segment
            if col_name in (d7_col, d2nd_col) and not disc and row_fill:
                cell.fill = row_fill

            # Action / Recommended bid — colour + bold
            if col_name in ('Action', 'Recommended bid') and row.get('Action') is not None:
                cell.fill = row_fill if row_fill else PatternFill('solid', start_color='D9E1F2')
                cell.font = Font(name='Arial', size=9, bold=True)

            # Daily cap — blue + bold
            if col_name == 'Daily Cap Suggestion' and val is not None:
                cell.fill = PatternFill('solid', start_color=FILLS['cap'])
                cell.font = Font(name='Arial', size=9, bold=True)

            # Number formats
            if col_name in ('fillRate', 'cvr', d7_col, d2nd_col):
                cell.number_format = '0.00%'
            elif col_name in ('spend', 'ecpp', 'ecpi', 'effectiveBidFloor',
                              'bidRate', 'Recommended bid', 'lowTier', 'midTier', 'highTier'):
                cell.number_format = '$#,##0.00'
            elif col_name in ('preloads', 'maxPreloads', 'installs', 'dailyCap'):
                cell.number_format = '#,##0'
            elif col_name in ('campaignId', 'siteId'):
                cell.number_format = '@'

    # Column widths
    col_widths = {
        'Key': 30, 'campaignId': 12, 'campaignName': 38, 'siteId': 10,
        'siteName': 45, 'status': 10, 'spend': 10, 'preloads': 10,
        'maxPreloads': 12, 'fillRate': 10, 'installs': 10, 'cvr': 8,
        'ecpp': 8, 'ecpi': 8, 'bidFloorGroupName': 20, 'effectiveBidFloor': 14,
        'bidRate': 10, 'dailyCap': 10, 'lowTier': 8, 'midTier': 8, 'highTier': 8,
        d7_col: 10, d2nd_col: 10, 'Action': 22, 'Recommended bid': 16,
        'Daily Cap Suggestion': 22
    }
    for col_idx, col_name in enumerate(out_cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 12)

    # No frozen panes
    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer


# ─── COLUMN INDEX HELPER ──────────────────────────────────────────────────────

def letter_to_index(letter):
    """Convert Excel column letter to 0-based index. E.g. 'I' → 8, 'K' → 10."""
    letter = letter.strip().upper()
    result = 0
    for char in letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1  # 0-based


if __name__ == '__main__':
    # Quick test
    print("optimizer.py loaded successfully")
    print(f"Column I = index {letter_to_index('I')}")
    print(f"Column K = index {letter_to_index('K')}")
